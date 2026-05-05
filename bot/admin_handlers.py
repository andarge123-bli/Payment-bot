"""
admin_handlers.py - Admin Interface (All text in Amharic)
Handles: Admin Management, Settings, User Management,
         Receipt Approval, Messaging, Financial Reports
"""

import os
import io
import logging
from datetime import datetime
from typing import Optional

from utils import (
    ETH_TZ,
    now_eth,
    to_ethiopian,
    eth_month_name,
    format_eth_date_storage,
    format_eth_datetime,
    prev_eth_months,
)

import pandas as pd
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputFile,
)
from telegram.ext import (
    ContextTypes,
    ConversationHandler,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters,
)

import database as db

logger = logging.getLogger(__name__)

CHANNEL_ID = os.getenv("CHANNEL_ID", "")

# ── Conversation States ─────────────────────────────────────────────────────
(
    ADD_ADMIN_ID,
    REMOVE_ADMIN_ID,
    EDIT_MSG_CHOOSE,
    EDIT_MSG_TEXT,
    SET_BILLING_START,
    SET_BILLING_END,
    ADD_BANK_NAME,
    ADD_BANK_ACCT,
    ADD_BANK_HOLDER,
    MANUAL_USER_ID,
    MANUAL_ACTION,
    MANUAL_NEW_NAME,
    REJECT_REASON,
    SUPPORT_REPLY_TEXT,
    BROADCAST_TEXT,
    BROADCAST_PHOTO,
    EDIT_USER_NAME_ID,
    EDIT_USER_NAME_VAL,
) = range(18)

EDITABLE_MESSAGES = {
    "msg_payment_start": "📢 የክፍያ ጊዜ ጅምር መልዕክት",
    "msg_reminder_one_day": "⏰ አንድ ቀን ቀረ ትዝታ",
    "msg_final_day": "🚨 የመጨረሻ ቀን መልዕክት",
    "msg_approved": "✅ ክፍያ ጸድቋል መልዕክት",
    "msg_rejected": "❌ ክፍያ ተቀባይነት አላገኘም መልዕክት",
}

NOTIFICATION_KEYS = {
    "notify_payment_start": "📢 የክፍያ ጊዜ ጅምር ማሳወቂያ",
    "notify_one_day": "⏰ አንድ ቀን ቀረ ማሳወቂያ",
    "notify_final_day": "🚨 የመጨረሻ ቀን ማሳወቂያ",
}


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────

def admin_required(func):
    """Decorator: Only allow admins to use this handler."""
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        if not db.is_admin(user_id) and not db.is_super_admin(user_id):
            await update.effective_message.reply_text("⛔ ይህን ትዕዛዝ ለመጠቀም ፈቃድ የለዎትም።")
            return ConversationHandler.END
        return await func(update, context)
    wrapper.__name__ = func.__name__
    return wrapper


def super_admin_required(func):
    """Decorator: Only super admin may proceed."""
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        if not db.is_super_admin(user_id):
            await update.effective_message.reply_text("⛔ ይህ ትዕዛዝ ለዋና አስተዳዳሪ ብቻ ነው።")
            return ConversationHandler.END
        return await func(update, context)
    wrapper.__name__ = func.__name__
    return wrapper


def _admin_main_menu_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🛡️ አስተዳዳሪ አስተዳደር", callback_data="adm_manage")],
        [InlineKeyboardButton("⚙️ ስርዓት ቅንብሮች", callback_data="adm_settings")],
        [InlineKeyboardButton("👥 ተጠቃሚ አስተዳደር", callback_data="adm_users")],
        [InlineKeyboardButton("📩 መልዕክቶች እና ደረሰኞች", callback_data="adm_inbox")],
        [InlineKeyboardButton("📊 የፋይናንስ ሪፖርት", callback_data="adm_report")],
    ])


# ─────────────────────────────────────────────
#  ADMIN MAIN PANEL
# ─────────────────────────────────────────────

@admin_required
async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.effective_message
    total = db.get_total_users_count()
    paid = db.get_total_paid_this_month()
    text = (
        "🔐 *የአስተዳዳሪ ፓነል*\n\n"
        f"👥 ጠቅላላ ተጠቃሚዎች: *{total}*\n"
        f"✅ ይሄ ወር ከፍለዋል: *{paid}*\n\n"
        "ከታቹ ምናሌ ምርጫዎን ያድርጉ:"
    )
    await msg.reply_text(text, reply_markup=_admin_main_menu_keyboard(), parse_mode="Markdown")


async def admin_panel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "adm_manage":
        await _show_admin_manage(query)
    elif data == "adm_settings":
        await _show_settings_menu(query)
    elif data == "adm_users":
        await _show_users_menu(query)
    elif data == "adm_inbox":
        await _show_inbox_menu(query)
    elif data == "adm_report":
        await _show_report_menu(query)
    elif data == "adm_back":
        total = db.get_total_users_count()
        paid = db.get_total_paid_this_month()
        text = (
            "🔐 *የአስተዳዳሪ ፓነል*\n\n"
            f"👥 ጠቅላላ ተጠቃሚዎች: *{total}*\n"
            f"✅ ይሄ ወር ከፍለዋል: *{paid}*\n\n"
            "ከታቹ ምናሌ ምርጫዎን ያድርጉ:"
        )
        await query.edit_message_text(text, reply_markup=_admin_main_menu_keyboard(), parse_mode="Markdown")


# ─────────────────────────────────────────────
#  🛡️ ADMIN MANAGEMENT
# ─────────────────────────────────────────────

async def _show_admin_manage(query):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ አስተዳዳሪ ጨምር", callback_data="adm_add_admin")],
        [InlineKeyboardButton("➖ አስተዳዳሪ አስወግድ", callback_data="adm_remove_admin")],
        [InlineKeyboardButton("📋 አስተዳዳሪዎች ዝርዝር", callback_data="adm_list_admins")],
        [InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_back")],
    ])
    await query.edit_message_text("🛡️ *አስተዳዳሪ አስተዳደር*\nምርጫዎን ያድርጉ:", reply_markup=kb, parse_mode="Markdown")


async def admin_manage_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    user_id = query.from_user.id

    if not db.is_super_admin(user_id):
        await query.answer("⛔ ለዋና አስተዳዳሪ ብቻ ነው።", show_alert=True)
        return

    if data == "adm_add_admin":
        await query.edit_message_text(
            "➕ *አስተዳዳሪ ጨምር*\n\nአዲሱን አስተዳዳሪ Telegram ID ያስገቡ:",
            parse_mode="Markdown",
        )
        return ADD_ADMIN_ID

    elif data == "adm_remove_admin":
        admins = db.get_all_admins()
        if not admins:
            await query.edit_message_text("❌ ምንም አስተዳዳሪ አልተገኘም።")
            return ConversationHandler.END
        kb_rows = []
        for a in admins:
            if not a.get("is_super"):
                kb_rows.append([InlineKeyboardButton(
                    f"🗑 ID: {a['telegram_id']}",
                    callback_data=f"remove_adm_{a['telegram_id']}"
                )])
        kb_rows.append([InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_manage")])
        await query.edit_message_text(
            "➖ *አስተዳዳሪ አስወግድ*\nማስወገድ የሚፈልጉትን ይምረጡ:",
            reply_markup=InlineKeyboardMarkup(kb_rows),
            parse_mode="Markdown",
        )

    elif data == "adm_list_admins":
        admins = db.get_all_admins()
        if not admins:
            text = "📋 አሁን ያሉ አስተዳዳሪዎች የሉም።"
        else:
            lines = ["📋 *አስተዳዳሪዎች ዝርዝር:*\n"]
            for i, a in enumerate(admins, 1):
                role = "⭐ ዋና" if a.get("is_super") else "🔑 ሁለተኛ"
                lines.append(f"{i}. ID: `{a['telegram_id']}` — {role}")
            text = "\n".join(lines)
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_manage")]])
        await query.edit_message_text(text, reply_markup=kb, parse_mode="Markdown")

    elif data.startswith("remove_adm_"):
        tid = int(data.split("_")[2])
        if db.remove_admin(tid):
            await query.edit_message_text(f"✅ ID `{tid}` አስተዳዳሪነት ተወግዷል።", parse_mode="Markdown")
        else:
            await query.edit_message_text("❌ ዋና አስተዳዳሪን ማስወገድ አይቻልም።")


async def receive_add_admin_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        new_admin_id = int(update.message.text.strip())
    except ValueError:
        await update.message.reply_text("❌ ትክክለኛ ID ያስገቡ (ቁጥር ብቻ):")
        return ADD_ADMIN_ID
    db.add_admin(new_admin_id, update.effective_user.id, is_super=False)
    await update.message.reply_text(f"✅ ID `{new_admin_id}` አስተዳዳሪ ሆኗል።", parse_mode="Markdown")
    return ConversationHandler.END


# ─────────────────────────────────────────────
#  ⚙️ SYSTEM SETTINGS
# ─────────────────────────────────────────────

async def _show_settings_menu(query):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✏️ መልዕክቶችን አርትዕ", callback_data="adm_edit_msgs")],
        [InlineKeyboardButton("🔔 ማሳወቂያ ቅንብር", callback_data="adm_notify_toggle")],
        [InlineKeyboardButton("📅 የክፍያ ዑደት", callback_data="adm_billing_cycle")],
        [InlineKeyboardButton("🏦 የባንክ ሒሳብ", callback_data="adm_bank")],
        [InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_back")],
    ])
    await query.edit_message_text("⚙️ *ስርዓት ቅንብሮች*\nምርጫዎን ያድርጉ:", reply_markup=kb, parse_mode="Markdown")


async def settings_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "adm_edit_msgs":
        rows = []
        for key, label in EDITABLE_MESSAGES.items():
            rows.append([InlineKeyboardButton(label, callback_data=f"edit_msg_{key}")])
        rows.append([InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_settings")])
        await query.edit_message_text(
            "✏️ *መልዕክቶችን አርትዕ*\nየትኛውን ማርትዕ ይፈልጋሉ?",
            reply_markup=InlineKeyboardMarkup(rows),
            parse_mode="Markdown",
        )

    elif data.startswith("edit_msg_"):
        key = data[len("edit_msg_"):]
        context.user_data["edit_msg_key"] = key
        current = db.get_setting(key, "")
        await query.edit_message_text(
            f"✏️ *{EDITABLE_MESSAGES.get(key, key)}*\n\n"
            f"አሁን ያለው:\n```{current}```\n\n"
            "አዲሱን ጽሑፍ ይላኩ (ለሰርዝ /cancel ይጠቀሙ):",
            parse_mode="Markdown",
        )
        return EDIT_MSG_TEXT

    elif data == "adm_notify_toggle":
        rows = []
        for key, label in NOTIFICATION_KEYS.items():
            current = db.get_setting(key, "true")
            status = "🟢 ነቅቷል" if current == "true" else "🔴 ጠፍቷል"
            rows.append([InlineKeyboardButton(f"{label} — {status}", callback_data=f"toggle_{key}")])
        rows.append([InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_settings")])
        await query.edit_message_text(
            "🔔 *ማሳወቂያ ቅንብሮች*\nለማብራት/ለማጥፋት ይምረጡ:",
            reply_markup=InlineKeyboardMarkup(rows),
            parse_mode="Markdown",
        )

    elif data.startswith("toggle_"):
        key = data[len("toggle_"):]
        current = db.get_setting(key, "true")
        new_val = "false" if current == "true" else "true"
        db.set_setting(key, new_val)
        status = "🟢 ነቅቷል" if new_val == "true" else "🔴 ጠፍቷል"
        await query.answer(f"{NOTIFICATION_KEYS.get(key, key)}: {status}", show_alert=True)
        # Refresh the toggle menu
        rows = []
        for k, label in NOTIFICATION_KEYS.items():
            cur = db.get_setting(k, "true")
            s = "🟢 ነቅቷል" if cur == "true" else "🔴 ጠፍቷል"
            rows.append([InlineKeyboardButton(f"{label} — {s}", callback_data=f"toggle_{k}")])
        rows.append([InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_settings")])
        await query.edit_message_text(
            "🔔 *ማሳወቂያ ቅንብሮች*:",
            reply_markup=InlineKeyboardMarkup(rows),
            parse_mode="Markdown",
        )

    elif data == "adm_billing_cycle":
        cycle = db.get_billing_cycle()
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("✏️ የጅምር ቀን ቀይር", callback_data="set_bill_start")],
            [InlineKeyboardButton("✏️ የማብቂያ ቀን ቀይር", callback_data="set_bill_end")],
            [InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_settings")],
        ])
        await query.edit_message_text(
            f"📅 *የክፍያ ዑደት*\n\n"
            f"የጅምር ቀን: **{cycle['start']}ኛ**\n"
            f"የማብቂያ ቀን: **{cycle['end']}ኛ**",
            reply_markup=kb,
            parse_mode="Markdown",
        )

    elif data == "set_bill_start":
        await query.edit_message_text("📅 አዲሱን የጅምር ቀን ቁጥር ያስገቡ (ለምሳሌ: 25):")
        return SET_BILLING_START

    elif data == "set_bill_end":
        await query.edit_message_text("📅 አዲሱን የማብቂያ ቀን ቁጥር ያስገቡ (ለምሳሌ: 5):")
        return SET_BILLING_END

    elif data == "adm_bank":
        accounts = db.get_active_bank_accounts()
        text = "🏦 *የባንክ ሒሳቦች*\n\n"
        if accounts:
            for a in accounts:
                text += (
                    f"🏦 {a['bank_name']}\n"
                    f"💳 {a['account_number']}\n"
                    f"👤 {a['account_holder']}\n\n"
                )
        else:
            text += "ምንም ሒሳብ አልተጨመረም።\n\n"
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("➕ ሒሳብ ጨምር", callback_data="bank_add")],
            [InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_settings")],
        ])
        await query.edit_message_text(text, reply_markup=kb, parse_mode="Markdown")

    elif data == "bank_add":
        await query.edit_message_text("🏦 የባንኩን ስም ያስገቡ (ለምሳሌ: ቢሮ፣ ዳሸን ወዘተ):")
        return ADD_BANK_NAME


async def receive_edit_msg_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    key = context.user_data.get("edit_msg_key")
    if not key:
        await update.message.reply_text("❌ ሂደቱ ጊዜ አልፎታል። እንደገና ይሞክሩ።")
        return ConversationHandler.END
    db.set_setting(key, update.message.text)
    await update.message.reply_text(f"✅ *{EDITABLE_MESSAGES.get(key, key)}* ዘምኗል!", parse_mode="Markdown")
    return ConversationHandler.END


async def receive_billing_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        day = int(update.message.text.strip())
        assert 1 <= day <= 31
    except (ValueError, AssertionError):
        await update.message.reply_text("❌ ትክክለኛ ቀን (1–31) ያስገቡ:")
        return SET_BILLING_START
    db.set_setting("billing_start_day", str(day))
    await update.message.reply_text(f"✅ የጅምር ቀን ወደ **{day}ኛ** ተቀይሯል።", parse_mode="Markdown")
    return ConversationHandler.END


async def receive_billing_end(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        day = int(update.message.text.strip())
        assert 1 <= day <= 31
    except (ValueError, AssertionError):
        await update.message.reply_text("❌ ትክክለኛ ቀን (1–31) ያስገቡ:")
        return SET_BILLING_END
    db.set_setting("billing_end_day", str(day))
    await update.message.reply_text(f"✅ የማብቂያ ቀን ወደ **{day}ኛ** ተቀይሯል።", parse_mode="Markdown")
    return ConversationHandler.END


async def receive_bank_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["new_bank_name"] = update.message.text.strip()
    await update.message.reply_text("💳 የሒሳብ ቁጥሩን ያስገቡ:")
    return ADD_BANK_ACCT


async def receive_bank_acct(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["new_bank_acct"] = update.message.text.strip()
    await update.message.reply_text("👤 የሒሳብ ባለቤቱን ሙሉ ስም ያስገቡ:")
    return ADD_BANK_HOLDER


async def receive_bank_holder(update: Update, context: ContextTypes.DEFAULT_TYPE):
    bank_name = context.user_data.get("new_bank_name", "")
    acct = context.user_data.get("new_bank_acct", "")
    holder = update.message.text.strip()
    db.add_bank_account(bank_name, acct, holder)
    await update.message.reply_text(
        f"✅ ሒሳብ ተጨምሯል!\n🏦 {bank_name}\n💳 {acct}\n👤 {holder}",
        parse_mode="Markdown",
    )
    return ConversationHandler.END


# ─────────────────────────────────────────────
#  👥 USER MANAGEMENT
# ─────────────────────────────────────────────

async def _show_users_menu(query):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 ሁሉም ተጠቃሚዎች", callback_data="users_all")],
        [InlineKeyboardButton("❌ ያልከፈሉ (ዕዳ ያለባቸው)", callback_data="users_debtors")],
        [InlineKeyboardButton("✏️ ተጠቃሚ ማስተካከያ", callback_data="users_manual")],
        [InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_back")],
    ])
    await query.edit_message_text("👥 *ተጠቃሚ አስተዳደር*:", reply_markup=kb, parse_mode="Markdown")


async def users_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "users_all":
        users = db.get_all_users()
        if not users:
            await query.edit_message_text("👥 ምንም ተጠቃሚ አልተገኘም።")
            return
        lines = [f"👥 *ሁሉም ተጠቃሚዎች ({len(users)}):*\n"]
        for u in users[:50]:  # Cap at 50 to avoid message length limit
            icon = "✅" if u["status"] == "paid" else "❌"
            lines.append(f"{icon} {u['name']} — `{u['telegram_id']}`")
        if len(users) > 50:
            lines.append(f"\n...እና {len(users)-50} ተጨማሪ ተጠቃሚዎች")
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_users")]])
        await query.edit_message_text("\n".join(lines), reply_markup=kb, parse_mode="Markdown")

    elif data == "users_debtors":
        debtors = db.get_unpaid_users()
        if not debtors:
            await query.edit_message_text("🎉 ሁሉም ተጠቃሚዎች ከፍለዋል!")
            return
        lines = [f"❌ *ያልከፈሉ ተጠቃሚዎች ({len(debtors)}):*\n"]
        for u in debtors[:50]:
            lines.append(f"• {u['name']} — `{u['telegram_id']}`")
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_users")]])
        await query.edit_message_text("\n".join(lines), reply_markup=kb, parse_mode="Markdown")

    elif data == "users_manual":
        await query.edit_message_text(
            "✏️ *ተጠቃሚ ማስተካከያ*\n\n"
            "የተጠቃሚው Telegram ID ያስገቡ:",
            parse_mode="Markdown",
        )
        return MANUAL_USER_ID


async def receive_manual_user_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        uid = int(update.message.text.strip())
    except ValueError:
        await update.message.reply_text("❌ ትክክለኛ ID ያስገቡ:")
        return MANUAL_USER_ID
    user = db.get_user(uid)
    if not user:
        await update.message.reply_text("❌ ይህ ተጠቃሚ አልተገኘም። ID እንደገና ያረጋግጡ:")
        return MANUAL_USER_ID
    context.user_data["manual_target_id"] = uid
    icon = "✅" if user["status"] == "paid" else "❌"
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ ተከፍሏል ምልክት", callback_data="manual_mark_paid")],
        [InlineKeyboardButton("❌ አልተከፈለም ምልክት", callback_data="manual_mark_unpaid")],
        [InlineKeyboardButton("✏️ ስም ቀይር", callback_data="manual_rename")],
    ])
    await update.message.reply_text(
        f"👤 *ተጠቃሚ:* {user['name']}\n"
        f"🆔 ID: `{uid}`\n"
        f"📊 ሁኔታ: {icon} {user['status']}\n\n"
        "ምን ማድረግ ይፈልጋሉ?",
        reply_markup=kb,
        parse_mode="Markdown",
    )
    return MANUAL_ACTION


async def manual_action_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    uid = context.user_data.get("manual_target_id")

    if data == "manual_mark_paid":
        db.update_user_status(uid, "paid")
        await query.edit_message_text(f"✅ ID `{uid}` ተከፍሏል ተብሎ ተመዝግቧል።", parse_mode="Markdown")
        return ConversationHandler.END

    elif data == "manual_mark_unpaid":
        db.update_user_status(uid, "unpaid")
        await query.edit_message_text(f"❌ ID `{uid}` አልተከፈለም ተብሎ ተለውጧል።", parse_mode="Markdown")
        return ConversationHandler.END

    elif data == "manual_rename":
        await query.edit_message_text("✏️ አዲሱን ስም ያስገቡ:")
        return MANUAL_NEW_NAME


async def receive_manual_new_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = context.user_data.get("manual_target_id")
    new_name = update.message.text.strip()
    db.update_user_name(uid, new_name)
    await update.message.reply_text(f"✅ ID `{uid}` ስም ወደ **{new_name}** ተቀይሯል።", parse_mode="Markdown")
    return ConversationHandler.END


# ─────────────────────────────────────────────
#  📩 INBOX (RECEIPTS + SUPPORT)
# ─────────────────────────────────────────────

async def _show_inbox_menu(query):
    pending = db.get_pending_payments()
    support = db.get_unanswered_support_messages()
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(f"📸 ደረሰኞች ({len(pending)})", callback_data="inbox_receipts")],
        [InlineKeyboardButton(f"💬 ድጋፍ ኢንቦክስ ({len(support)})", callback_data="inbox_support")],
        [InlineKeyboardButton("📢 ሁሉም ተጠቃሚዎች ላክ", callback_data="inbox_broadcast")],
        [InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_back")],
    ])
    await query.edit_message_text("📩 *መልዕክቶች እና ደረሰኞች*:", reply_markup=kb, parse_mode="Markdown")


async def inbox_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "inbox_receipts":
        payments = db.get_pending_payments()
        if not payments:
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_inbox")]])
            await query.edit_message_text("✅ አሁን ምንም ያልታዩ ደረሰኞች የሉም።", reply_markup=kb)
            return
        await query.edit_message_text(
            f"📸 *{len(payments)} ደረሰኞች ቀርበዋል*\n\nለእያንዳንዱ ደረሰኝ ይፈርዱ:"
        )
        for p in payments:
            user = db.get_user(p["telegram_id"])
            name = user["name"] if user else "ያልታወቀ"
            channel_link = f"https://t.me/c/{str(CHANNEL_ID).replace('-100', '')}/{p['receipt_channel_msg_id']}"
            kb = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("✅ ጸድቋል", callback_data=f"approve_{p['id']}"),
                    InlineKeyboardButton("❌ ተቀባይነት የለም", callback_data=f"reject_{p['id']}"),
                ]
            ])
            # Show Ethiopian payment date if stored, otherwise convert from Gregorian
            eth_date_str = p.get("eth_payment_date", "")
            if not eth_date_str:
                try:
                    gd = datetime.strptime(str(p.get("created_at", ""))[:10], "%Y-%m-%d")
                    eth_date_str = format_eth_date_storage(gd)
                except Exception:
                    eth_date_str = f"{p['year']}-{p['month']:02d}"

            text = (
                f"👤 *{name}*\n"
                f"🆔 ID: `{p['telegram_id']}`\n"
                f"📅 ቀን (ዓ.ም): {eth_date_str}\n"
                f"🔗 [ደረሰኝ ይመልከቱ]({channel_link})"
            )
            await query.message.reply_text(text, reply_markup=kb, parse_mode="Markdown")

    elif data.startswith("approve_"):
        payment_id = int(data.split("_")[1])
        payment = db.get_payment_by_id(payment_id)
        if payment and db.approve_payment(payment_id):
            user = db.get_user(payment["telegram_id"])
            name = user["name"] if user else "ያልታወቀ"
            # Use Ethiopian month name from the payment record
            eth_month_num = payment.get("month", 0)
            month_str = eth_month_name(eth_month_num) if eth_month_num else eth_month_name(to_ethiopian(now_eth())[1])
            msg_template = db.get_setting("msg_approved", "✅ ክፍያዎ ጸድቋል!")
            msg = msg_template.format(name=name, month=month_str)
            try:
                await query.get_bot().send_message(chat_id=payment["telegram_id"], text=msg)
            except Exception as e:
                logger.warning(f"Could not notify user {payment['telegram_id']}: {e}")
            await query.edit_message_text(f"✅ ID `{payment['telegram_id']}` ክፍያ ጸድቋል!", parse_mode="Markdown")
        else:
            await query.edit_message_text("❌ ሂደቱ አልተሳካም።")

    elif data.startswith("reject_"):
        payment_id = int(data.split("_")[1])
        context.user_data["reject_payment_id"] = payment_id
        await query.edit_message_text("❌ *ክፍያ ለምን ተቀባይነት አላገኘም?*\n\nምክንያቱን ያስገቡ:", parse_mode="Markdown")
        return REJECT_REASON

    elif data == "inbox_support":
        messages = db.get_unanswered_support_messages()
        if not messages:
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_inbox")]])
            await query.edit_message_text("✅ ምንም ያልተመለሰ ጥያቄ የለም።", reply_markup=kb)
            return
        await query.edit_message_text(f"💬 *{len(messages)} ጥያቄዎች:*")
        for m in messages:
            user = db.get_user(m["telegram_id"])
            name = user["name"] if user else "ያልታወቀ"
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("↩️ መልስ ላክ", callback_data=f"reply_sup_{m['id']}")]
            ])
            await query.message.reply_text(
                f"👤 *{name}* (`{m['telegram_id']}`)\n"
                f"📩 *መልዕክት:* {m['message']}\n"
                f"🕐 {m['created_at'][:16]}",
                reply_markup=kb,
                parse_mode="Markdown",
            )

    elif data.startswith("reply_sup_"):
        msg_id = int(data.split("_")[2])
        context.user_data["support_reply_id"] = msg_id
        await query.edit_message_text("↩️ *መልስ ላክ*\n\nምላሽዎን ያስገቡ:", parse_mode="Markdown")
        return SUPPORT_REPLY_TEXT

    elif data == "inbox_broadcast":
        await query.edit_message_text(
            "📢 *ሁሉም ተጠቃሚዎች ላክ*\n\n"
            "መልዕክቱን ወይም ፎቶ ጽሑፍ ያስገቡ (ወይም /cancel ለመሰረዝ):",
            parse_mode="Markdown",
        )
        return BROADCAST_TEXT


async def receive_reject_reason(update: Update, context: ContextTypes.DEFAULT_TYPE):
    reason = update.message.text.strip()
    payment_id = context.user_data.get("reject_payment_id")
    if not payment_id:
        await update.message.reply_text("❌ ሂደቱ ጊዜ አልፎታል።")
        return ConversationHandler.END
    payment = db.get_payment_by_id(payment_id)
    db.reject_payment(payment_id, reason)
    if payment:
        msg_template = db.get_setting("msg_rejected", "❌ ክፍያዎ ተቀባይነት አላገኘም።\nምክንያት: {reason}")
        msg = msg_template.format(reason=reason)
        try:
            await update.get_bot().send_message(chat_id=payment["telegram_id"], text=msg)
        except Exception as e:
            logger.warning(f"Could not notify user: {e}")
    await update.message.reply_text("✅ ተጠቃሚው ውሳኔው ተሳወቀ።")
    return ConversationHandler.END


async def receive_support_reply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg_id = context.user_data.get("support_reply_id")
    reply_text = update.message.text.strip()
    admin_id = update.effective_user.id
    support_msg = db.get_support_message_by_id(msg_id)
    if not support_msg:
        await update.message.reply_text("❌ ጥያቄው አልተገኘም።")
        return ConversationHandler.END
    db.reply_to_support_message(msg_id, reply_text, admin_id)
    try:
        await update.get_bot().send_message(
            chat_id=support_msg["telegram_id"],
            text=f"📩 *ከድጋፍ ቡድን:*\n\n{reply_text}",
            parse_mode="Markdown",
        )
    except Exception as e:
        logger.warning(f"Could not send support reply: {e}")
    await update.message.reply_text("✅ መልሱ ተልኳል!")
    return ConversationHandler.END


async def receive_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    users = db.get_all_users()
    success, failed = 0, 0
    msg = update.message

    for user in users:
        try:
            if msg.photo:
                await msg.get_bot().send_photo(
                    chat_id=user["telegram_id"],
                    photo=msg.photo[-1].file_id,
                    caption=msg.caption or "",
                )
            else:
                await msg.get_bot().send_message(
                    chat_id=user["telegram_id"],
                    text=f"📢 *ማሳወቂያ:*\n\n{msg.text}",
                    parse_mode="Markdown",
                )
            success += 1
        except Exception:
            failed += 1

    await update.message.reply_text(
        f"📢 *ስርጭት ተጠናቅቋል!*\n\n✅ ተልኮ: {success}\n❌ አልተሳካም: {failed}",
        parse_mode="Markdown",
    )
    return ConversationHandler.END


# ─────────────────────────────────────────────
#  📊 FINANCIAL REPORT
# ─────────────────────────────────────────────

def _month_picker_keyboard(prefix: str) -> InlineKeyboardMarkup:
    """Build an inline keyboard showing the last 6 Ethiopian months."""
    months = prev_eth_months(6)
    buttons = []
    for i in range(0, len(months), 2):
        row = []
        for y, m in months[i:i + 2]:
            label = f"{eth_month_name(m)} {y}"
            row.append(InlineKeyboardButton(label, callback_data=f"{prefix}_{y}_{m}"))
        buttons.append(row)
    buttons.append([InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_report")])
    return InlineKeyboardMarkup(buttons)


async def _show_report_menu(query):
    now_dt = now_eth()
    eth_yr, eth_mo, _ = to_ethiopian(now_dt)
    total = db.get_total_users_count()
    paid = db.get_total_paid_this_month()
    unpaid = total - paid
    paid_pct = round(paid / total * 100) if total > 0 else 0
    filled = round(paid_pct / 10)
    bar = "█" * filled + "░" * (10 - filled)
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📥 የክፍያ ሪፖርት (Excel)", callback_data="report_excel_pick")],
        [InlineKeyboardButton("📋 የተገኝነት ሪፖርት (Excel)", callback_data="report_attend_pick")],
        [InlineKeyboardButton("📣 ያልከፈሉ ተጠቃሚዎች ያሳውቁ", callback_data="report_notify_pick")],
        [InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_back")],
    ])
    text = (
        f"📊 *የፋይናንስ ዳሽቦርድ — {eth_month_name(eth_mo)} {eth_yr} (ዓ.ም)*\n\n"
        f"👥 ጠቅላላ ተጠቃሚዎች: *{total}*\n"
        f"✅ ተከፍሏል:          *{paid}*\n"
        f"❌ አልተከፈለም:       *{unpaid}*\n\n"
        f"📈 የክፍያ መጠን: *{paid_pct}%*\n"
        f"`[{bar}]`"
    )
    await query.edit_message_text(text, reply_markup=kb, parse_mode="Markdown")


async def _generate_payment_excel(query, eth_mo: int, eth_yr: int):
    """Build and send the payment Excel for the given Ethiopian month/year."""
    await query.edit_message_text("⏳ Excel ሪፖርት እየተዘጋጀ ነው...")
    payments = db.get_monthly_payments(eth_mo, eth_yr)
    users_map = {u["telegram_id"]: u for u in db.get_all_users()}

    rows = []
    for p in payments:
        user = users_map.get(p["telegram_id"], {})
        eth_date = p.get("eth_payment_date", "") or f"{p['year']}-{p['month']:02d}"
        rows.append({
            "ስም": user.get("name", "ያልታወቀ"),
            "Telegram ID": p["telegram_id"],
            "ወር (ዓ.ም)": f"{eth_month_name(p['month'])} {p['year']}",
            "ቀን (ዓ.ም)": eth_date,
            "ሁኔታ": "ተከፍሏል" if p["status"] == "approved" else p["status"],
        })

    all_users = db.get_all_users()
    paid_ids = {p["telegram_id"] for p in payments}
    for u in all_users:
        if u["telegram_id"] not in paid_ids:
            rows.append({
                "ስም": u["name"],
                "Telegram ID": u["telegram_id"],
                "ወር (ዓ.ም)": f"{eth_month_name(eth_mo)} {eth_yr}",
                "ቀን (ዓ.ም)": "",
                "ሁኔታ": "አልተከፈለም",
            })

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="ሪፖርት")
        ws = writer.sheets["ሪፖርት"]
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 40)
    output.seek(0)

    filename = f"report_{eth_yr}_{eth_mo:02d}.xlsx"
    await query.message.reply_document(
        document=InputFile(output, filename=filename),
        caption=f"📊 *{eth_month_name(eth_mo)} {eth_yr} (ዓ.ም) ሪፖርት*",
        parse_mode="Markdown",
    )


async def _generate_attendance_excel(query, eth_mo: int, eth_yr: int):
    """Build and send the attendance Excel for the given Ethiopian month/year."""
    await query.edit_message_text("⏳ የተገኝነት ሪፖርት እየተዘጋጀ ነው...")
    rows = db.get_attendance_data(eth_mo, eth_yr)

    if not rows:
        await query.edit_message_text("❌ ምንም ተጠቃሚ አልተገኘም።")
        return

    month_label = eth_month_name(eth_mo)

    on_time  = sum(1 for r in rows if r["Timeliness"] == "On Time")
    late     = sum(1 for r in rows if r["Timeliness"] == "Late")
    unpaid   = sum(1 for r in rows if r["Timeliness"] == "Unpaid")
    total    = len(rows)
    paid_pct = round((on_time + late) / total * 100) if total > 0 else 0

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook

    wb = Workbook()

    # ════ Sheet 1: Full Attendance Table ════════════════════════════
    ws1 = wb.active
    ws1.title = "የተገኝነት ዝርዝር"

    GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
    YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
    RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BODY_FONT   = Font(size=10)
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws1.merge_cells("A1:H1")
    title_cell = ws1["A1"]
    title_cell.value = f"የተገኝነት ሪፖርት — {month_label} {eth_yr} (ዓ.ም)"
    title_cell.font = Font(bold=True, size=14, color="1F3864")
    title_cell.alignment = CENTER
    ws1.row_dimensions[1].height = 28

    headers    = ["ተ.ቁ", "ስም", "Telegram ID", "ወር", "ሁኔታ", "ወቅታዊነት", "የክፍያ ቀን", "ማስታወሻ"]
    col_widths = [6, 24, 16, 14, 14, 14, 18, 20]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=2, column=col_idx, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    ws1.row_dimensions[2].height = 20

    display_cols = ["ተ.ቁ", "ስም", "Telegram ID", "ወር", "ሁኔታ", "ወቅታዊነት", "የክፍያ ቀን"]
    for row_idx, row_data in enumerate(rows, 3):
        timeliness = row_data.get("Timeliness", "Unpaid")
        if timeliness == "On Time":
            fill = GREEN_FILL
            note = "ቀደም ብሎ ከፍሏል"
        elif timeliness == "Late":
            fill = YELLOW_FILL
            note = "ዘግይቶ ከፍሏል"
        else:
            fill = RED_FILL
            note = "አልከፈለም"

        for col_idx, col_key in enumerate(display_cols, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=row_data.get(col_key, ""))
            cell.fill      = fill
            cell.font      = BODY_FONT
            cell.border    = THIN_BORDER
            cell.alignment = CENTER if col_idx != 2 else LEFT

        note_cell = ws1.cell(row=row_idx, column=8, value=note)
        note_cell.fill      = fill
        note_cell.font      = Font(size=10, italic=True)
        note_cell.border    = THIN_BORDER
        note_cell.alignment = CENTER

    ws1.freeze_panes = "A3"

    # ════ Sheet 2: Summary Dashboard ════════════════════════════════
    ws2 = wb.create_sheet(title="ማጠቃለያ")

    GOLD_FILL    = PatternFill("solid", fgColor="FFD700")
    SECTION_FONT = Font(bold=True, size=12, color="1F3864")
    VALUE_FONT   = Font(bold=True, size=20, color="1F3864")

    def _summary_block(ws, start_row, label, value, fill, pct=None):
        ws.merge_cells(start_row=start_row, start_column=2,
                       end_row=start_row,   end_column=4)
        lc = ws.cell(row=start_row, column=2, value=label)
        lc.font = SECTION_FONT; lc.fill = fill
        lc.alignment = LEFT;    lc.border = THIN_BORDER

        vc = ws.cell(row=start_row, column=5, value=value)
        vc.font = VALUE_FONT; vc.fill = fill
        vc.alignment = CENTER; vc.border = THIN_BORDER

        if pct is not None:
            pc = ws.cell(row=start_row, column=6, value=f"{pct}%")
            pc.font = Font(bold=True, size=12, color="555555")
            pc.fill = fill; pc.alignment = CENTER; pc.border = THIN_BORDER

        ws.row_dimensions[start_row].height = 36

    for col, w in [("A", 3), ("B", 6), ("C", 20), ("D", 10), ("E", 12), ("F", 10)]:
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("B1:F1")
    t = ws2["B1"]
    t.value     = f"ማጠቃለያ — {month_label} {eth_yr} (ዓ.ም)"
    t.font      = Font(bold=True, size=16, color="1F3864")
    t.alignment = CENTER
    ws2.row_dimensions[1].height = 32
    ws2.row_dimensions[2].height = 10

    _summary_block(ws2, 3, "👥 ጠቅላላ ተጠቃሚዎች",    total,   PatternFill("solid", fgColor="DDEBF7"))
    _summary_block(ws2, 4, "✅ ወቅቱን ጠብቀው ከፍለዋል", on_time, GREEN_FILL,
                   round(on_time / total * 100) if total else 0)
    _summary_block(ws2, 5, "⏰ ዘግይተው ከፍለዋል",      late,    YELLOW_FILL,
                   round(late / total * 100) if total else 0)
    _summary_block(ws2, 6, "❌ አልከፈሉም",            unpaid,  RED_FILL,
                   round(unpaid / total * 100) if total else 0)
    _summary_block(ws2, 7, "📈 ጠቅላላ የክፍያ መጠን",   f"{paid_pct}%", GOLD_FILL)

    ws2.row_dimensions[8].height = 14
    legend_items = [
        (9,  GREEN_FILL,  "🟢 ወቅቱን ጠብቆ — ሁለት ቀን ወይም ከዚያ በፊት ከፍሏል"),
        (10, YELLOW_FILL, "🟡 ዘግይቶ — ከመጨረሻ አንድ-ሁለት ቀን ጠብቆ ከፍሏል"),
        (11, RED_FILL,    "🔴 አልከፈለም — ምንም ክፍያ አልቀረበም"),
    ]
    for r, fill, leg_text in legend_items:
        ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws2.cell(row=r, column=2, value=leg_text)
        c.fill = fill; c.font = Font(size=10, italic=True)
        c.alignment = LEFT; c.border = THIN_BORDER
        ws2.row_dimensions[r].height = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    cycle = db.get_billing_cycle()
    filename = f"attendance_{eth_yr}_{eth_mo:02d}.xlsx"
    caption = (
        f"📋 *የተገኝነት ሪፖርት — {month_label} {eth_yr} (ዓ.ም)*\n\n"
        f"👥 ጠቅላላ: *{total}*   ✅ ወቅቱን: *{on_time}*   "
        f"⏰ ዘግይቶ: *{late}*   ❌ አልከፈለም: *{unpaid}*\n"
        f"📈 የክፍያ መጠን: *{paid_pct}%*\n\n"
        f"_የዑደት ቀናት: {cycle['start']}ኛ — {cycle['end']}ኛ_"
    )
    await query.message.reply_document(
        document=InputFile(output, filename=filename),
        caption=caption,
        parse_mode="Markdown",
    )


async def report_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "adm_report":
        await _show_report_menu(query)

    elif data == "report_excel_pick":
        await query.edit_message_text(
            "📥 *የክፍያ ሪፖርት — ወር ይምረጡ:*",
            reply_markup=_month_picker_keyboard("report_excel"),
            parse_mode="Markdown",
        )

    elif data == "report_attend_pick":
        await query.edit_message_text(
            "📋 *የተገኝነት ሪፖርት — ወር ይምረጡ:*",
            reply_markup=_month_picker_keyboard("report_attend"),
            parse_mode="Markdown",
        )

    elif data.startswith("report_excel_") and data.count("_") == 3:
        _, _, yr, mo = data.split("_")
        await _generate_payment_excel(query, int(mo), int(yr))

    elif data.startswith("report_attend_") and data.count("_") == 3:
        _, _, yr, mo = data.split("_")
        await _generate_attendance_excel(query, int(mo), int(yr))

    elif data == "report_notify_pick":
        await query.edit_message_text(
            "📣 *ያሳወቅ — ወር ይምረጡ:*\n\n"
            "ለምን ወር ያልከፈሉ ተጠቃሚዎች ማሳወቅ ይፈልጋሉ?",
            reply_markup=_month_picker_keyboard("report_nfy"),
            parse_mode="Markdown",
        )

    elif data.startswith("report_nfy_") and not data.startswith("report_nfyok_"):
        # format: report_nfy_{year}_{month}  — show preview + confirm
        parts = data.split("_")
        yr, mo = int(parts[2]), int(parts[3])
        unpaid_users = db.get_unpaid_users_for_month(mo, yr)
        count = len(unpaid_users)
        month_label = eth_month_name(mo)

        if count == 0:
            await query.edit_message_text(
                f"✅ *{month_label} {yr} (ዓ.ም)*\n\n"
                "ሁሉም ተጠቃሚዎች ለዚህ ወር ከፍለዋል! ምንም ማሳወቂያ አያስፈልግም።",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("◀️ ተመለስ", callback_data="adm_report")]
                ]),
            )
            return

        # Show preview of first 5 names
        preview_lines = [f"  • {u['name']}" for u in unpaid_users[:5]]
        more = f"\n  _...እና {count - 5} ተጨማሪ_" if count > 5 else ""
        preview = "\n".join(preview_lines) + more

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(
                f"📨 አዎ — {count} ሰዎች ላክ",
                callback_data=f"report_nfyok_{yr}_{mo}",
            )],
            [InlineKeyboardButton("❌ ሰርዝ", callback_data="adm_report")],
        ])
        await query.edit_message_text(
            f"📣 *{month_label} {yr} (ዓ.ም) — ያሳወቅ ቅድሚያ ዕይታ*\n\n"
            f"❌ *{count} ተጠቃሚዎች* ለዚህ ወር ገና አልከፈሉም:\n\n"
            f"{preview}\n\n"
            "ማሳወቂያ ወደ ሁሉም ልካቸው?",
            reply_markup=kb,
            parse_mode="Markdown",
        )

    elif data.startswith("report_nfyok_"):
        # format: report_nfyok_{year}_{month}  — send messages
        parts = data.split("_")
        yr, mo = int(parts[2]), int(parts[3])
        month_label = eth_month_name(mo)
        unpaid_users = db.get_unpaid_users_for_month(mo, yr)

        if not unpaid_users:
            await query.edit_message_text("✅ ሁሉም ተከፍሏል — ምንም ለማሳወቅ የለም።")
            return

        await query.edit_message_text(
            f"⏳ ማሳወቂያ ወደ {len(unpaid_users)} ተጠቃሚዎች እየተላከ ነው..."
        )

        cycle = db.get_billing_cycle()
        template = db.get_setting(
            "msg_final_day",
            "⚠️ ትዝታ! ለ{month} ወር ክፍያዎን ገና አልፈጸሙም። እባክዎ ወዲያው ይፈጽሙ!",
        )
        sent = 0
        failed = 0
        for user in unpaid_users:
            try:
                msg = template.format(
                    name=user["name"],
                    month=month_label,
                    end_day=cycle["end"],
                    start_day=cycle["start"],
                )
                await query.get_bot().send_message(
                    chat_id=user["telegram_id"],
                    text=f"📣 *ትዝታ — {month_label} {yr} (ዓ.ም)*\n\n{msg}",
                    parse_mode="Markdown",
                )
                sent += 1
            except Exception as e:
                logger.warning(f"Notify failed for {user['telegram_id']}: {e}")
                failed += 1

        result_text = (
            f"✅ *ማሳወቂያ ተልኳል!*\n\n"
            f"📅 ወር: *{month_label} {yr} (ዓ.ም)*\n"
            f"📨 ተልኳል:   *{sent}*\n"
        )
        if failed:
            result_text += f"❌ አልተላከም: *{failed}*\n"
        await query.edit_message_text(
            result_text,
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("◀️ ወደ ሪፖርት ተመለስ", callback_data="adm_report")]
            ]),
        )


# ─────────────────────────────────────────────
#  AUTOMATED REMINDER JOBS (called by JobQueue)
# ─────────────────────────────────────────────

async def send_payment_start_reminder(context: ContextTypes.DEFAULT_TYPE):
    if db.get_setting("notify_payment_start", "true") != "true":
        return
    cycle = db.get_billing_cycle()
    template = db.get_setting("msg_payment_start", "📢 የክፍያ ጊዜ ደርሷል!")
    msg = template.format(start_day=cycle["start"], end_day=cycle["end"])
    users = db.get_unpaid_users()
    # Reset all users to unpaid at cycle start
    db.reset_all_users_to_unpaid()
    for user in users:
        try:
            await context.bot.send_message(chat_id=user["telegram_id"], text=msg)
        except Exception as e:
            logger.warning(f"Reminder failed for {user['telegram_id']}: {e}")


async def send_one_day_reminder(context: ContextTypes.DEFAULT_TYPE):
    if db.get_setting("notify_one_day", "true") != "true":
        return
    cycle = db.get_billing_cycle()
    template = db.get_setting("msg_reminder_one_day", "⚠️ ነገ የክፍያ ቀን ነው!")
    msg = template.format(end_day=cycle["end"])
    for user in db.get_unpaid_users():
        try:
            await context.bot.send_message(chat_id=user["telegram_id"], text=msg)
        except Exception as e:
            logger.warning(f"One-day reminder failed for {user['telegram_id']}: {e}")


async def send_final_day_reminder(context: ContextTypes.DEFAULT_TYPE):
    if db.get_setting("notify_final_day", "true") != "true":
        return
    cycle = db.get_billing_cycle()
    template = db.get_setting("msg_final_day", "🚨 ዛሬ የመጨረሻ ቀን ነው!")
    msg = template.format(end_day=cycle["end"])
    for user in db.get_unpaid_users():
        try:
            await context.bot.send_message(chat_id=user["telegram_id"], text=msg)
        except Exception as e:
            logger.warning(f"Final day reminder failed for {user['telegram_id']}: {e}")


async def monthly_cycle_reset_job(context: ContextTypes.DEFAULT_TYPE):
    """
    Runs the day after the billing end day at 00:05 Ethiopian time.
    1. Captures a full summary of the closing cycle (before reset).
    2. Resets ALL users to 'unpaid' for the new cycle.
    3. Sends a detailed Amharic summary report to all admins.
    """
    now = now_eth()
    cycle = db.get_billing_cycle()
    end_day = cycle["end"]

    # Determine which Ethiopian month/year just closed
    eth_year, eth_month, eth_day = to_ethiopian(now)

    # If today is day 1, the closing cycle was the previous Ethiopian month
    if eth_day == 1:
        closing_eth_month = eth_month - 1 if eth_month > 1 else 13
        closing_eth_year = eth_year if eth_month > 1 else eth_year - 1
    else:
        closing_eth_month = eth_month
        closing_eth_year = eth_year

    logger.info(f"🔄 Monthly reset job: closing cycle {eth_month_name(closing_eth_month)} {closing_eth_year} (ዓ.ም)")

    # ── Step 1: Capture summary BEFORE reset ────────────────────────────────
    summary = db.get_cycle_summary(closing_eth_month, closing_eth_year)

    # ── Step 2: Reset all users to unpaid ───────────────────────────────────
    db.reset_all_users_to_unpaid()
    logger.info(f"✅ All {summary['total_users']} users reset to 'unpaid' for new cycle.")

    # ── Step 3: Build Amharic report message ────────────────────────────────
    paid_pct = (
        round(summary["total_paid"] / summary["total_users"] * 100)
        if summary["total_users"] > 0 else 0
    )

    # Payment rate bar (10-block visual)
    filled = round(paid_pct / 10)
    bar = "█" * filled + "░" * (10 - filled)

    # List up to 10 still-unpaid users
    unpaid_lines = []
    for u in summary["unpaid_users"][:10]:
        unpaid_lines.append(f"  • {u['name']} (`{u['telegram_id']}`)")
    unpaid_preview = "\n".join(unpaid_lines) if unpaid_lines else "  _ሁሉም ተጠቃሚዎች ከፍለዋል!_ 🎉"
    more_unpaid = (
        f"\n  _...እና {len(summary['unpaid_users']) - 10} ተጨማሪ ሰዎች_"
        if len(summary["unpaid_users"]) > 10 else ""
    )

    eth_month_label = eth_month_name(closing_eth_month)
    report = (
        f"📊 *የወር ዑደት ሪፖርት — {eth_month_label} {closing_eth_year} (ዓ.ም)*\n"
        f"━━━━━━━━━━━━━━━━━━\n\n"
        f"👥 ጠቅላላ ተጠቃሚዎች:   *{summary['total_users']}*\n"
        f"✅ ክፍያ ፈጽመዋል:      *{summary['total_paid']}*\n"
        f"❌ አልከፈሉም:          *{summary['total_unpaid']}*\n"
        f"⏳ በጥበቃ ላይ:         *{summary['total_pending']}*\n"
        f"🚫 ተቀባይነት አላገኘም:  *{summary['total_rejected']}*\n\n"
        f"📈 የክፍያ መጠን: *{paid_pct}%*\n"
        f"`[{bar}]`\n\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"❌ *ያልከፈሉ ተጠቃሚዎች:*\n"
        f"{unpaid_preview}{more_unpaid}\n\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"🔄 *ሁሉም ተጠቃሚዎች ለአዲሱ ዑደት 'አልተከፈለም' ሆነዋል።*\n"
        f"🕐 ዳግም ማስጀመሪያ ቀን: {format_eth_datetime(now)} (ዓ.ም)"
    )

    # ── Step 4: Notify all admins ────────────────────────────────────────────
    admins = db.get_all_admins()
    super_id = int(os.getenv("ADMIN_ID", "0"))

    # Always include super admin even if not in DB
    admin_ids = {a["telegram_id"] for a in admins}
    if super_id:
        admin_ids.add(super_id)

    notified = 0
    for admin_id in admin_ids:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=report,
                parse_mode="Markdown",
            )
            notified += 1
        except Exception as e:
            logger.warning(f"Could not send monthly report to admin {admin_id}: {e}")

    logger.info(f"✅ Monthly reset report sent to {notified} admin(s).")


# ─────────────────────────────────────────────
#  CANCEL HANDLER
# ─────────────────────────────────────────────

async def cancel_conv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ ሂደቱ ተሰርዟል።")
    return ConversationHandler.END


# ─────────────────────────────────────────────
#  CONVERSATION HANDLER BUILDER
# ─────────────────────────────────────────────

def build_admin_conversation() -> ConversationHandler:
    """Build and return the master admin ConversationHandler."""
    return ConversationHandler(
        entry_points=[
            CommandHandler("admin", admin_panel),
            CallbackQueryHandler(admin_manage_callback, pattern=r"^adm_(add_admin|remove_admin|list_admins|manage)$|^remove_adm_"),
            CallbackQueryHandler(settings_callback, pattern=r"^adm_(edit_msgs|notify_toggle|billing_cycle|bank)|^edit_msg_|^toggle_|^set_bill_|^bank_"),
            CallbackQueryHandler(users_callback, pattern=r"^users_"),
            CallbackQueryHandler(inbox_callback, pattern=r"^inbox_|^approve_|^reject_|^reply_sup_"),
            CallbackQueryHandler(report_callback, pattern=r"^report_"),
            CallbackQueryHandler(admin_panel_callback, pattern=r"^adm_(settings|users|inbox|report|back)$"),
        ],
        states={
            ADD_ADMIN_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_add_admin_id)],
            EDIT_MSG_TEXT: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_edit_msg_text)],
            SET_BILLING_START: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_billing_start)],
            SET_BILLING_END: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_billing_end)],
            ADD_BANK_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_bank_name)],
            ADD_BANK_ACCT: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_bank_acct)],
            ADD_BANK_HOLDER: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_bank_holder)],
            MANUAL_USER_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_manual_user_id)],
            MANUAL_ACTION: [CallbackQueryHandler(manual_action_callback, pattern=r"^manual_")],
            MANUAL_NEW_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_manual_new_name)],
            REJECT_REASON: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_reject_reason)],
            SUPPORT_REPLY_TEXT: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_support_reply)],
            BROADCAST_TEXT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_broadcast),
                MessageHandler(filters.PHOTO, receive_broadcast),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel_conv)],
        allow_reentry=True,
        per_message=False,
    )
